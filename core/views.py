from django.contrib import messages
from functools import wraps
from django.db.models.functions import TruncDate, Coalesce
from django.db.models import ExpressionWrapper, F

from django.db.models import Sum

from django.db.models.fields import DateTimeField

import face_recognition
import os
import base64
from django.http import JsonResponse
from django.shortcuts import render, redirect,get_object_or_404
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from .models import ApplicationName,ApplicationType,Department
from django.http import HttpResponse
from django.contrib.auth.decorators import user_passes_test, login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.db import transaction
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from AccountApp.models import AppUser, Attendance, Program, Semester, Department
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import pickle
# Custom Decorators

def is_admin(user):
    return user.is_authenticated and user.is_superuser

def check_attendance(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        # Your custom logic here
        if request.user.is_authenticated:
            user = request.user
            current_date = datetime.now().date()

            try:
                attendance_instance = Attendance.objects.get(user=user, date=current_date)

                if attendance_instance.entry_time and attendance_instance.exit_time:
                    # Both entry and exit are present, redirect to 'core:StopMessage'
                    return redirect('core:StopMessageView')

            except Attendance.DoesNotExist:
                # Attendance instance not present for the day, allow access to the view
                return view_func(request, *args, **kwargs)

        # If the user is not authenticated, redirect to a login page
        else:
            return redirect('AccountApp:custom_login')  # Replace 'login' with your actual login URL

        # If the control reaches here, it means the user is authenticated but attendance conditions are not met
        return view_func(request, *args, **kwargs)

    return wrapper
#----------------------------------------------------------------------------------------------------------

def SetProjectNameView(request):
    try:
        set_name = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_name = "No Name"

    # If the user is already logged in, redirect to a different page
    if request.user.is_authenticated:
        return redirect('core:index')

    context = {'Name_of_Project': set_name}
    return render(request, 'core/welcome.html', context)


@login_required(login_url="AccountApp:custom_login")
def IndexPageView(request):
    get_user = request.user
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    context = {'Set_Logo': set_logo, 'user_info': get_user}
    return render(request, 'core/index.html', context)


def add_padding(data):
    # Ensure that the Base64 data has proper padding
    padding = len(data) % 4
    if padding:
        data += '=' * (4 - padding)
    return data
@csrf_exempt
@user_passes_test(is_admin, login_url="AccountApp:custom_login")

def capture_image(request):

    get_user = request.user
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    if request.method == 'POST' and 'image_data' in request.POST:
        get_id = request.POST.get('user_id')
        try:
            # Retrieve the AppUser based on the custom unique ID
            user_data = AppUser.objects.get(custom_unique_id=get_id)
            # Get the media root from Django settings
            media_root = settings.MEDIA_ROOT
            # Generate a timestamp for the file name
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            folder_name = f"{get_id}_{user_data.first_name}_{user_data.last_name}"
            # Create the folder path using the custom unique ID and names
            folder_path = os.path.join(media_root + "/dataset", folder_name)
            # Create the file name with the timestamp
            file_name = f"{user_data.first_name}_{user_data.last_name}_{timestamp}.png"
            try:
                # Check if the folder already exists, and create it if not
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                    print(f"Folder '{folder_name}' created successfully at '{folder_path}'.")
                else:
                    print(f"Folder '{folder_name}' already exists at '{folder_path}'.")
            except Exception as e:
                print(f"Error creating folder: {e}")
            # Process the captured image data
            image_data = request.POST['image_data'].split(',')[1]
            image_content = ContentFile(base64.b64decode(image_data), name=file_name)
            # Create the full image path
            image_path = os.path.join(folder_path, file_name)
            # Save the image file to the specified path
            with open(os.path.join(settings.MEDIA_ROOT, image_path), 'wb') as img_file:
                img_file.write(image_content.read())
            # Update the image path in the AppUser model
            #user_data.image_path = image_path
            user_data.save()
            return JsonResponse({'message': 'Image captured and saved successfully!'})
        except AppUser.DoesNotExist:
            # Handle the case where the user with the given custom unique ID does not exist
            return JsonResponse({'error': 'User not found'}, status=404)
    else:
        get_data = AppUser.objects.filter(is_superuser=False)
        print(get_data)
        context = {'user_data': get_data, 'user_info':get_user, 'Set_Logo':set_logo}
        return render(request, 'core/capture_image.html', context)

def train_system(dataset_path):

    known_face_encodings = []
    known_face_names = []
    for person_folder in os.listdir(dataset_path):
        person_path = os.path.join(dataset_path, person_folder)
        if os.path.isdir(person_path):
            folder_parts = person_folder.split('_')
            if len(folder_parts) == 3:
                person_id, first_name, last_name = folder_parts[0], folder_parts[1], folder_parts[2]
                person_name = f"{first_name} {last_name}"
            for image_file in os.listdir(person_path):
                image_path = os.path.join(person_path, image_file)
                print(person_path)
                print("")
                print(image_file)
                # Load the image
                image = face_recognition.load_image_file(image_path)
                # Get the face encoding
                face_encoding = face_recognition.face_encodings(image)
                if len(face_encoding) > 0:
                    # Store the face encoding and the corresponding person's name
                    known_face_encodings.append(face_encoding[0])
                    known_face_names.append({'id': person_id, 'name': person_name})
    trained_data_folder_name = "trained_data"
    dat_file_folder = os.path.join(settings.BASE_DIR, trained_data_folder_name)
    if not os.path.exists(dat_file_folder):
        os.makedirs(dat_file_folder)
    print(dat_file_folder)
    # Save the trained data
    with open(os.path.join(dat_file_folder, 'trained_data.dat'), 'wb') as file:
        data = {'encodings': known_face_encodings, 'names': known_face_names}
        pickle.dump(data, file)


@user_passes_test(is_admin, login_url="AccountApp:custom_login")
def TrainOnDataView(request):
    get_user = request.user
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    context = {}
    if request.method == 'POST':
        datasetfolder_name = "dataset"
        dataset_path = os.path.join(settings.MEDIA_ROOT, datasetfolder_name)
        train_system(dataset_path)
        context = {'Message':'Training completed on the active Dataset', 'user_info':get_user, 'Set_Logo':set_logo}
        return render(request, 'core/train_data.html', context)
    context = {'user_info':get_user, 'Set_Logo':set_logo}
    return render(request, 'core/train_data.html', context)

@csrf_exempt
@login_required(login_url="AccountApp:custom_login")
@check_attendance
def DetectPersonView(request):
    get_user = request.user
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    context = {'message': None}

    if request.method == 'POST' and 'image_data' in request.POST:
        try:
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            capture_image_folder = "detect_image"
            image_recognize_path = os.path.join(settings.BASE_DIR, capture_image_folder)

            if not os.path.exists(image_recognize_path):
                os.makedirs(image_recognize_path)

            file_name = f"_{timestamp}.png"
            image_data = request.POST['image_data'].split(',')[1]
            image_content = ContentFile(base64.b64decode(image_data), name=file_name)
            image_path = os.path.join(image_recognize_path, file_name)

            with open(image_path, 'wb') as img_file:
                img_file.write(image_content.read())

            get_trained_data_file = os.path.join(settings.BASE_DIR, "trained_data")
            complete_path = os.path.join(get_trained_data_file, "trained_data.dat")

            with open(complete_path, 'rb') as file:
                data = pickle.load(file)
                known_face_encodings = data['encodings']
                known_face_names = data['names']

            unknown_image = face_recognition.load_image_file(image_path)
            face_locations = face_recognition.face_locations(unknown_image)
            face_encodings = face_recognition.face_encodings(unknown_image, face_locations)

            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
                person_info = {'id': "Unknown", 'name': "Unknown"}

                if True in matches:
                    first_match_index = matches.index(True)
                    person_info = known_face_names[first_match_index]
                    person_id = person_info['id']
                    person_name = person_info['name']
                    print(f"Person ID: {person_id}, Person Name: {person_name}, Location: {top},{right},{bottom},{left}")

                    # Check if the user has already marked attendance on the same day
                    user = AppUser.objects.get(custom_unique_id=person_id)
                    current_date = datetime.now().date()

                    with transaction.atomic():
                        attendance_instance = Attendance.objects.filter(user=user, date=current_date).first()

                        if attendance_instance:
                            # User has already marked attendance, consider it as exit attendance
                            if not attendance_instance.exit_time:
                                attendance_instance.exit_time = datetime.now().time()
                                attendance_instance.save()
                                context['message'] = 'Exit time marked successfully!'
                            else:
                                context['message'] = 'Attendance already marked for entry and exit.'
                        else:
                            # User is marking attendance for the first time, consider it as entry attendance
                            new_attendance = Attendance(
                                user=user,
                                date=current_date,
                                entry_time=datetime.now().time(),
                                is_present = True
                            )
                            new_attendance.save()
                            context['message'] = 'Entry time marked successfully.'

                    print("Attendance Marked Successfully")
                    os.remove(image_path)
                    return JsonResponse({'success': True, 'message': context['message']})
        except Exception as e:
            os.remove(image_path)
            return HttpResponse('Some Error Occurred')
            os.remove(image_path)
    context = {'user_info':get_user, 'Set_Logo':set_logo}
    return render(request, 'core/detect_person.html', context)

def ViewStopMessage(request):
    get_user = request.user
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    context = {'user_info':get_user, 'Set_Logo':set_logo}
    return render(request, 'core/stop.html', context)

def AttendanceSuccessView(request):
    get_user = request.user
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    context = {'user_info':get_user, 'Set_Logo':set_logo}
    return render(request, 'core/success_attendance.html', context)
def AttendanceFailedView(request):
    get_user = request.user
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    context = {'user_info':get_user, 'Set_Logo':set_logo}
    return render(request, 'core/failed_attendance.html', context)
@login_required(login_url="AccountApp:custom_login")
def AttendenceSearchView(request):
    get_user = request.user
    application_type = ApplicationType.objects.first()

    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None

    get_deptt_name = Department.objects.all()
    get_program_name = Program.objects.all()
    get_semester_name = Semester.objects.all()

    users_in_department = AppUser.objects.filter(is_superuser=False)

    # Retrieve search parameters from the URL
    select_deptt_name = request.GET.get('dept_name')
    select_program_name = request.GET.get('program_name')
    select_semester_name = request.GET.get('semester_name')

    if select_deptt_name:
        get_deptt_object = Department.objects.get(id=select_deptt_name)
        users_in_department = users_in_department.filter(department=get_deptt_object)

    if select_program_name:
        get_program_object = Program.objects.get(id=select_program_name)
        users_in_department = users_in_department.filter(program=get_program_object)

    if select_semester_name:
        get_semester_object = Semester.objects.get(id=select_semester_name)
        users_in_department = users_in_department.filter(semester=get_semester_object)

    # Add pagination
    paginator = Paginator(users_in_department, 10)  # Show 10 users per page
    page = request.GET.get('page')
    try:
        users_in_department = paginator.page(page)
    except PageNotAnInteger:
        users_in_department = paginator.page(1)
    except EmptyPage:
        users_in_department = paginator.page(paginator.num_pages)

    context = {'Department_Names': get_deptt_name, 'Program_Names': get_program_name,
               'Semester_Names': get_semester_name, 'Users_Info': users_in_department,
               'user_info': get_user, 'Set_Logo': set_logo, 'App_type':application_type}

    return render(request, 'core/attendence_search.html', context)

@login_required(login_url="AccountApp:custom_login")
def AttendenceRecordView(request, pk):
    get_user = get_object_or_404(AppUser, custom_unique_id=pk)
    get_attendance_record = Attendance.objects.filter(user=get_user).order_by('-date', '-entry_time')
    # Check if export button is clicked
    if 'export' in request.GET:
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        # Define header row
        header = ['Unique ID', 'User Name', 'Date', 'Time', 'Present']
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header_text
            ws[f'{col_letter}1'].font = Font(bold=True)
        # Populate data rows
        for row_num, attendance in enumerate(get_attendance_record, 2):
            ws.cell(row=row_num, column=1, value=get_user.custom_unique_id)
            ws.cell(row=row_num, column=2, value=get_user.username)
            ws.cell(row=row_num, column=3, value=attendance.date)
            ws.cell(row=row_num, column=4, value=attendance.time)
            ws.cell(row=row_num, column=5, value=attendance.is_present)
        # Create a response object with the appropriate headers
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={get_user.custom_unique_id}_attendance_record.xlsx'
        # Save the workbook to the response
        wb.save(response)
        return response
     # Add pagination
    paginator = Paginator(get_attendance_record, 10)  # Show 10 records per page
    page = request.GET.get('page')
    try:
        attendance_records = paginator.page(page)
    except PageNotAnInteger:
        attendance_records = paginator.page(1)
    except EmptyPage:
        attendance_records = paginator.page(paginator.num_pages)
    context = {'Attendance_Record': attendance_records, 'User': get_user}
    return render(request, 'core/person_record.html', context)

def UpdateRecordView(request, pk):
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    person_record = get_object_or_404(AppUser, id=pk)
    get_department = Department.objects.all()
    get_program = Program.objects.all()
    get_semester = Semester.objects.all()
    current_loggedin_user = request.user
    user_role = 'superuser' if person_record.is_superuser else 'teacher' if person_record.is_teacher else 'other'
    application_type = ApplicationType.objects.first()

    if request.method == 'POST':
        Unique_ID = request.POST.get('Unique_ID')
        username = request.POST.get('username')
        fname = request.POST.get('fname')
        lname = request.POST.get('lname')
        email = request.POST.get('email')
        department_id = request.POST.get('department')  # Get department ID from the form

        # Validate that department_id is a valid integer
        try:
            department_id = int(department_id)
        except ValueError:
            # Handle the case where department_id is not a valid integer
            department_id = None

        role = request.POST.get('role')

        if application_type.type_app == 'EDU':
            program_id = request.POST.get('program')  # Get program ID from the form
            semester_id = request.POST.get('semester')  # Get semester ID from the form

            # Validate that program_id and semester_id are valid integers
            try:
                program_id = int(program_id)
                semester_id = int(semester_id)
            except ValueError:
                # Handle the case where program_id or semester_id is not a valid integer
                program_id = None
                semester_id = None
        else:
            program_id = None
            semester_id = None

        # Get Department, Program, and Semester instances or set to None if IDs are not valid
        department_instance = get_object_or_404(Department, id=department_id) if department_id else None
        program_instance = get_object_or_404(Program, id=program_id) if program_id else None
        semester_instance = get_object_or_404(Semester, id=semester_id) if semester_id else None

        # Update the AppUser instance
        update_user = AppUser.objects.get(id=pk)
        update_user.username = username
        update_user.first_name = fname
        update_user.last_name = lname
        update_user.email = email
        update_user.department = department_instance
        update_user.role = role

        if application_type.type_app == 'EDU':
            update_user.program = program_instance
            update_user.semester = semester_instance

        update_user.save()
        return redirect('core:ViewAttendenceSearch')

    context = {'Record': person_record, 'dept_names': get_department, 'user_role': user_role,
               'loggedin_user': current_loggedin_user, 'programs': get_program, 'Student_Semester': get_semester,
               'app_type': application_type, 'Set_Logo': set_logo, 'user_info':current_loggedin_user}
    return render(request, 'core/update_record.html', context)

def DeleteRecordView(request, pk):
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    get_user = AppUser.objects.get(id=pk)

    if request.method == 'POST':
        if 'yes' in request.POST:
            get_user.delete()
            return redirect('core:ViewAttendenceSearch')
        else:
            return redirect('core:ViewAttendenceSearch')
    context = {'user_data':get_user, 'Set_Logo': set_logo, 'user_info':request.user}
    return render(request, 'core/delete_record.html', context)

def calculate_hours_per_month(request):
    get_user = request.user
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    current_month = timezone.now().month
    current_year = timezone.now().year

    # Get the logged-in user's attendance for the current month
    user_attendance = Attendance.objects.filter(
        user=request.user,
        date__month=current_month,
        date__year=current_year
    )

    total_seconds = 0

    for attendance in user_attendance:
        attendance.hours_worked = attendance.calculate_hours_worked()
        total_seconds += attendance.hours_worked[0] * 3600 + attendance.hours_worked[1] * 60 + attendance.hours_worked[2]

    total_hours = total_seconds // 3600
    total_minutes = (total_seconds % 3600) // 60
    total_seconds_remainder = total_seconds % 60

    average_hour = (total_hours + total_minutes / 60 + total_seconds_remainder / 3600) / 240

    context = {
            'user_attendance': user_attendance,
            'total_hours': average_hour,
            'user_info':get_user, 'Set_Logo':set_logo
        }

    return render(request, 'core/hours_calculation.html', context)


def AdminAverageCalView(request):
    get_user_info = request.user
    get_user = None
    try:
        set_logo = ApplicationName.objects.latest('id')
    except ApplicationName.DoesNotExist:
        set_logo = None
    context = {'user_info':get_user_info, 'Set_Logo':set_logo}
    if request.method ==   'POST':
        get_user_custom_id = request.POST.get('user_custom_id')
        # Get the current month and year
        current_month = timezone.now().month
        current_year = timezone.now().year

        get_user = AppUser.objects.get(custom_unique_id = get_user_custom_id)
        user_attendance = Attendance.objects.filter(
            user=get_user,
            date__month=current_month,
            date__year=current_year
        )

        total_seconds = 0

        for attendance in user_attendance:
            attendance.hours_worked = attendance.calculate_hours_worked()
            total_seconds += attendance.hours_worked[0] * 3600 + attendance.hours_worked[1] * 60 + attendance.hours_worked[2]

        total_hours = total_seconds // 3600
        total_minutes = (total_seconds % 3600) // 60
        total_seconds_remainder = total_seconds % 60

        average_hour = (total_hours + total_minutes / 60 + total_seconds_remainder / 3600) / 240

        context = {
            'user_attendance': user_attendance,
            'total_hours': average_hour,
            'user_info':get_user_info, 'Set_Logo':set_logo
        }

    return render(request, 'core/admin_hours_calculation.html', context)